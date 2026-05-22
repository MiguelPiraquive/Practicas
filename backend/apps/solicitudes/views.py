from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse, FileResponse
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Solicitud, TipoDocumentoSolicitado, Parentesco, TipoDocumentoIdentidad
from .serializers import (
    SolicitudSerializer,
    TipoDocumentoSolicitadoSerializer,
    ParentescoSerializer,
    TipoDocumentoIdentidadSerializer,
)
from . import hosvital, envio, openwa
from apps.bitacora.utils import registrar_log
from apps.permisos.permissions import PermisosPorAccionMixin


class _CatalogoSoftDeleteMixin:
    """Mixin reutilizable para catálogos editables con soft-delete y bitácora."""

    bitacora_modelo = "Catalogo"

    def perform_create(self, serializer):
        instance = serializer.save()
        registrar_log(
            usuario=self.request.user,
            accion=f"Crear {self.bitacora_modelo}",
            modelo=self.bitacora_modelo,
            registro_id=instance.id,
            detalle=f"Creado '{instance}'",
        )

    def perform_update(self, serializer):
        instance = serializer.save()
        registrar_log(
            usuario=self.request.user,
            accion=f"Actualizar {self.bitacora_modelo}",
            modelo=self.bitacora_modelo,
            registro_id=instance.id,
            detalle=f"Actualizado '{instance}'",
        )

    def perform_destroy(self, instance):
        registro = str(instance)
        instance_id = instance.id
        instance.activo = False
        instance.save(update_fields=["activo"])
        registrar_log(
            usuario=self.request.user,
            accion=f"Desactivar {self.bitacora_modelo}",
            modelo=self.bitacora_modelo,
            registro_id=instance_id,
            detalle=f"Desactivado '{registro}'",
        )


class TipoDocumentoSolicitadoViewSet(_CatalogoSoftDeleteMixin, PermisosPorAccionMixin, viewsets.ModelViewSet):
    queryset = TipoDocumentoSolicitado.objects.all()
    serializer_class = TipoDocumentoSolicitadoSerializer
    search_fields = ["nombre", "descripcion"]
    filterset_fields = ["activo"]
    ordering_fields = ["nombre", "fecha_creacion"]
    bitacora_modelo = "TipoDocumentoSolicitado"
    permisos_requeridos = {
        "list":           "tipos_documento_solicitado.ver",
        "retrieve":       "tipos_documento_solicitado.ver",
        "create":         "tipos_documento_solicitado.crear",
        "update":         "tipos_documento_solicitado.editar",
        "partial_update": "tipos_documento_solicitado.editar",
        "destroy":        "tipos_documento_solicitado.eliminar",
    }


class ParentescoViewSet(_CatalogoSoftDeleteMixin, PermisosPorAccionMixin, viewsets.ModelViewSet):
    queryset = Parentesco.objects.all()
    serializer_class = ParentescoSerializer
    search_fields = ["nombre", "descripcion"]
    filterset_fields = ["activo"]
    ordering_fields = ["nombre", "fecha_creacion"]
    bitacora_modelo = "Parentesco"
    permisos_requeridos = {
        "list":           "parentescos.ver",
        "retrieve":       "parentescos.ver",
        "create":         "parentescos.crear",
        "update":         "parentescos.editar",
        "partial_update": "parentescos.editar",
        "destroy":        "parentescos.eliminar",
    }


class TipoDocumentoIdentidadViewSet(_CatalogoSoftDeleteMixin, PermisosPorAccionMixin, viewsets.ModelViewSet):
    queryset = TipoDocumentoIdentidad.objects.all()
    serializer_class = TipoDocumentoIdentidadSerializer
    search_fields = ["codigo", "nombre"]
    filterset_fields = ["activo"]
    ordering_fields = ["codigo", "nombre", "fecha_creacion"]
    bitacora_modelo = "TipoDocumentoIdentidad"
    permisos_requeridos = {
        "list":           "tipos_doc_identidad.ver",
        "retrieve":       "tipos_doc_identidad.ver",
        "create":         "tipos_doc_identidad.crear",
        "update":         "tipos_doc_identidad.editar",
        "partial_update": "tipos_doc_identidad.editar",
        "destroy":        "tipos_doc_identidad.eliminar",
    }


class SolicitudViewSet(PermisosPorAccionMixin, viewsets.ModelViewSet):
    queryset = Solicitud.objects.select_related(
        "paciente", "solicitado_por", "responsable_busqueda",
        "funcionario_entrega",
    ).all()
    serializer_class = SolicitudSerializer
    filterset_fields = ["estado", "paciente__numero_documento"]
    search_fields = [
        "paciente__nombres",
        "paciente__apellidos",
        "paciente__numero_documento",
    ]
    ordering_fields = ["fecha_solicitud", "estado"]
    permisos_requeridos = {
        "list":                 "solicitudes.ver",
        "retrieve":             "solicitudes.ver",
        "create":               "solicitudes.crear",
        "update":               "solicitudes.editar",
        "partial_update":       "solicitudes.editar",
        "destroy":              "solicitudes.eliminar",
        "cambiar_estado":       "solicitudes.cambiar_estado",
        "motivos_historicos":   "solicitudes.ver",
        "estadisticas":         "solicitudes.ver",
        "exportar":             "solicitudes.exportar",
        "generar_hc":           "solicitudes.generar_hc",
        "subir_hc":             "solicitudes.generar_hc",
        "eliminar_hc":          "solicitudes.generar_hc",
        "descargar_hc":         "solicitudes.ver",
        "enviar_hc":            "solicitudes.enviar_hc",
        "whatsapp_estado":      "solicitudes.enviar_hc",
        "whatsapp_refrescar":   "solicitudes.enviar_hc",
    }

    @action(detail=False, methods=["get"], url_path="whatsapp-estado")
    def whatsapp_estado(self, request):
        """Diagnóstico del gateway WhatsApp (OpenWA): qué sesión está activa, etc."""
        listo, info = openwa.estado_sesion()
        return Response(info, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="whatsapp-refrescar")
    def whatsapp_refrescar(self, request):
        """Fuerza re-detección de la sesión activa (invalida el caché)."""
        sid, fuente = openwa.refrescar_sesion()
        return Response(
            {"session_id": sid, "fuente": fuente, "ok": bool(sid)},
            status=status.HTTP_200_OK,
        )

    def perform_create(self, serializer):
        solicitud = serializer.save(solicitado_por=self.request.user)
        # Registrar la hora de recibo automáticamente al crear.
        if not solicitud.hora_recibo:
            solicitud.hora_recibo = timezone.localtime().time().replace(microsecond=0)
        # El usuario que registra la solicitud asume inmediatamente la búsqueda,
        # por lo que la solicitud entra directamente en estado EN_BUSQUEDA.
        if solicitud.estado == "SOLICITADA":
            solicitud.estado = "EN_BUSQUEDA"
            if not solicitud.responsable_busqueda_id:
                solicitud.responsable_busqueda = self.request.user
        solicitud.save(update_fields=["hora_recibo", "estado", "responsable_busqueda"])
        registrar_log(
            usuario=self.request.user,
            accion="Crear solicitud",
            modelo="Solicitud",
            registro_id=solicitud.id,
            detalle=f"Solicitud creada para paciente {solicitud.paciente}",
        )

    @action(detail=True, methods=["patch"], url_path="cambiar-estado")
    def cambiar_estado(self, request, pk=None):
        solicitud = self.get_object()
        nuevo_estado = request.data.get("estado")

        transiciones_validas = {
            "SOLICITADA": ["EN_BUSQUEDA"],
            "EN_BUSQUEDA": ["LISTA"],
            "LISTA": ["ENTREGADA"],
            "ENTREGADA": [],
        }

        estados_permitidos = transiciones_validas.get(solicitud.estado, [])
        if nuevo_estado not in estados_permitidos:
            return Response(
                {"error": f"No se puede pasar de '{solicitud.estado}' a '{nuevo_estado}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        estado_anterior = solicitud.estado
        solicitud.estado = nuevo_estado

        if nuevo_estado == "EN_BUSQUEDA":
            solicitud.responsable_busqueda = request.user
        elif nuevo_estado == "LISTA":
            solicitud.fecha_lista = timezone.now()
        elif nuevo_estado == "ENTREGADA":
            solicitud.fecha_entrega = timezone.now()
            solicitud.entregado_a = request.data.get("entregado_a", "")
            solicitud.funcionario_entrega = request.user
            # Hora de envío real al entregar
            if not solicitud.hora_envio:
                solicitud.hora_envio = timezone.localtime().time().replace(microsecond=0)
            # Medios de entrega (se pueden actualizar al entregar)
            if "medio_entrega_fisico" in request.data:
                solicitud.medio_entrega_fisico = request.data["medio_entrega_fisico"]
            if "medio_entrega_correo" in request.data:
                solicitud.medio_entrega_correo = request.data["medio_entrega_correo"]
            if "medio_entrega_whatsapp" in request.data:
                solicitud.medio_entrega_whatsapp = request.data["medio_entrega_whatsapp"]

        solicitud.save()

        registrar_log(
            usuario=request.user,
            accion=f"Cambiar estado: {estado_anterior} -> {nuevo_estado}",
            modelo="Solicitud",
            registro_id=solicitud.id,
            detalle=f"Solicitud SOL-{solicitud.id} cambio de {estado_anterior} a {nuevo_estado}",
        )

        serializer = self.get_serializer(solicitud)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="motivos-historicos")
    def motivos_historicos(self, request):
        """Devuelve la lista de motivos únicos usados anteriormente (para autocompletar)."""
        motivos = (
            Solicitud.objects.exclude(motivo_solicitud="")
            .values_list("motivo_solicitud", flat=True)
            .distinct()
            .order_by("motivo_solicitud")
        )
        return Response(list(motivos))

    @action(detail=False, methods=["get"], url_path="estadisticas")
    def estadisticas(self, request):
        """Devuelve conteos por estado y total, aplicando filtros opcionales de fecha."""
        from django.db.models import Count

        qs = self.filter_queryset(self.get_queryset())
        fecha_desde = request.query_params.get("fecha_desde")
        fecha_hasta = request.query_params.get("fecha_hasta")
        if fecha_desde:
            qs = qs.filter(fecha_solicitud__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha_solicitud__date__lte=fecha_hasta)

        conteos = dict(qs.values_list("estado").annotate(c=Count("id")))
        return Response({
            "total": qs.count(),
            "solicitada": conteos.get("SOLICITADA", 0),
            "en_busqueda": conteos.get("EN_BUSQUEDA", 0),
            "lista": conteos.get("LISTA", 0),
            "entregada": conteos.get("ENTREGADA", 0),
        })

    @action(detail=False, methods=["get"], url_path="exportar")
    def exportar(self, request):
        solicitudes = self.filter_queryset(self.get_queryset())
        fecha_desde = request.query_params.get("fecha_desde")
        fecha_hasta = request.query_params.get("fecha_hasta")
        if fecha_desde:
            solicitudes = solicitudes.filter(fecha_solicitud__date__gte=fecha_desde)
        if fecha_hasta:
            solicitudes = solicitudes.filter(fecha_solicitud__date__lte=fecha_hasta)
        solicitudes = solicitudes.order_by("fecha_solicitud")

        wb = Workbook()
        ws = wb.active
        ws.title = "ENTREGA HISTORIAS CLINICAS"

        # Encabezados exactos del Excel de ventanilla
        headers = [
            "FECHA",
            "HORA ENVÍO",
            "TIPO DOC PACIENTE",
            "No. DOC PACIENTE",
            "NOMBRE DEL PACIENTE",
            "DOCUMENTO SOLICITADO",
            "MOTIVO DE LA SOLICITUD",
            "FECHAS DE LA(S) ATENCIÓN(ES)",
            "TIPO DE SOLICITUD - TRÁMITE",
            "TIENE AUTORIZADO",
            "NOMBRE DEL AUTORIZADO",
            "PARENTESCO",
            "TIPO DOC AUTORIZADO",
            "No. DOC AUTORIZADO",
            "FUNCIONARIO QUE ENTREGA",
            "MEDIO ENTREGA - FÍSICO",
            "MEDIO ENTREGA - CORREO",
            "MEDIO ENTREGA - WHATSAPP",
            "FECHA ENTREGA",
            "OBSERVACIÓN",
        ]

        # Estilo de encabezado
        header_fill = PatternFill(start_color="1F5C8B", end_color="1F5C8B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Filas de datos
        for sol in solicitudes:
            pac = sol.paciente
            ws.append([
                sol.fecha_solicitud.strftime("%Y-%m-%d") if sol.fecha_solicitud else "",
                sol.hora_envio.strftime("%H:%M") if sol.hora_envio else
                    (sol.fecha_solicitud.strftime("%H:%M") if sol.fecha_solicitud else ""),
                pac.tipo_documento,
                pac.numero_documento,
                f"{pac.apellidos} {pac.nombres}".strip(),
                sol.get_documento_solicitado_display() if sol.documento_solicitado else "",
                sol.motivo_solicitud or sol.motivo,
                sol.fechas_atencion,
                sol.get_tipo_tramite_display() if sol.tipo_tramite else "",
                "SÍ" if sol.tiene_autorizado else "NO",
                sol.nombre_autorizado,
                sol.parentesco_autorizado,
                sol.tipo_doc_autorizado,
                sol.numero_doc_autorizado,
                sol.funcionario_entrega.nombre_completo if sol.funcionario_entrega else
                    (sol.responsable_busqueda.nombre_completo if sol.responsable_busqueda else ""),
                "X" if sol.medio_entrega_fisico else "",
                "X" if sol.medio_entrega_correo else "",
                "X" if sol.medio_entrega_whatsapp else "",
                sol.fecha_entrega.strftime("%Y-%m-%d %H:%M") if sol.fecha_entrega else "",
                sol.observaciones,
            ])

        # Ajuste de ancho de columnas
        col_widths = [12, 10, 8, 16, 30, 20, 22, 22, 18, 8, 28, 14, 8, 16, 22, 8, 8, 8, 18, 35]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="ENTREGA_HC_VENTANILLA_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        )
        wb.save(response)
        return response

    # =====================================================================
    # HOSVITAL — descarga del PDF de la HC desde el sistema hospitalario
    # =====================================================================
    @action(detail=True, methods=["post"], url_path="generar-hc")
    def generar_hc(self, request, pk=None):
        """Llama a Hosvital y guarda el PDF en archivo_hc."""
        solicitud = self.get_object()
        pac = solicitud.paciente

        from datetime import datetime as _dt

        def _parse(d):
            if not d:
                return None
            if hasattr(d, "year"):
                return d
            try:
                return _dt.strptime(str(d), "%Y-%m-%d").date()
            except (ValueError, TypeError):
                return None

        fecha_desde = _parse(request.data.get("fecha_desde") or solicitud.fecha_atencion_desde)
        fecha_hasta = _parse(request.data.get("fecha_hasta") or solicitud.fecha_atencion_hasta)
        tipo_hc = request.data.get("tipo_hc") or solicitud.documento_solicitado or "HC_COMPLETA"

        solicitud.hosvital_estado = "EN_PROCESO"
        solicitud.save(update_fields=["hosvital_estado"])

        try:
            pdf_bytes, mensaje = hosvital.descargar_historia_clinica(
                tipo_documento_paciente=pac.tipo_documento,
                numero_documento_paciente=pac.numero_documento,
                tipo_hc=tipo_hc,
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
                nombre_paciente=f"{pac.nombres} {pac.apellidos}".strip(),
            )
        except Exception as e:
            solicitud.hosvital_estado = "ERROR"
            solicitud.hosvital_mensaje = str(e)
            solicitud.hosvital_fecha = timezone.now()
            solicitud.save(update_fields=["hosvital_estado", "hosvital_mensaje", "hosvital_fecha"])
            registrar_log(
                usuario=request.user, accion="Hosvital ERROR",
                modelo="Solicitud", registro_id=solicitud.id, detalle=str(e),
            )
            return Response({"error": str(e)}, status=status.HTTP_502_BAD_GATEWAY)

        nombre_archivo = f"SOL-{solicitud.id:04d}_{pac.numero_documento}.pdf"
        solicitud.archivo_hc.save(nombre_archivo, ContentFile(pdf_bytes), save=False)
        solicitud.hosvital_estado = "RECIBIDA"
        solicitud.hosvital_mensaje = mensaje
        solicitud.hosvital_fecha = timezone.now()
        solicitud.save()

        registrar_log(
            usuario=request.user, accion="Generar HC desde Hosvital",
            modelo="Solicitud", registro_id=solicitud.id, detalle=mensaje,
        )
        return Response(self.get_serializer(solicitud).data)

    @action(detail=True, methods=["post"], url_path="subir-hc")
    def subir_hc(self, request, pk=None):
        """Recibe el PDF generado externamente por Hosvital y lo guarda."""
        solicitud = self.get_object()
        archivo = request.FILES.get("archivo") or request.FILES.get("file")
        if not archivo:
            return Response(
                {"error": "No se recibió ningún archivo. Adjunte el PDF de la HC."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        nombre_original = (archivo.name or "").lower()
        content_type = (archivo.content_type or "").lower()
        if not (nombre_original.endswith(".pdf") or content_type == "application/pdf"):
            return Response(
                {"error": "El archivo debe ser un PDF."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        max_bytes = 25 * 1024 * 1024  # 25 MB
        if archivo.size and archivo.size > max_bytes:
            return Response(
                {"error": "El archivo supera el tamaño máximo permitido (25 MB)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        pac = solicitud.paciente
        nombre_archivo = f"SOL-{solicitud.id:04d}_{pac.numero_documento}.pdf"
        solicitud.archivo_hc.save(nombre_archivo, archivo, save=False)
        solicitud.hosvital_estado = "RECIBIDA"
        solicitud.hosvital_mensaje = f"PDF cargado manualmente ({archivo.name})"
        solicitud.hosvital_fecha = timezone.now()
        solicitud.save()

        registrar_log(
            usuario=request.user, accion="Subir HC manual",
            modelo="Solicitud", registro_id=solicitud.id,
            detalle=f"Archivo {archivo.name} ({archivo.size} bytes)",
        )
        return Response(self.get_serializer(solicitud).data)

    @action(detail=True, methods=["delete"], url_path="eliminar-hc")
    def eliminar_hc(self, request, pk=None):
        """Elimina el PDF de HC cargado para poder subir uno nuevo."""
        solicitud = self.get_object()
        if not solicitud.archivo_hc:
            return Response(
                {"error": "No hay archivo de HC para eliminar."},
                status=status.HTTP_404_NOT_FOUND,
            )
        nombre = solicitud.archivo_hc.name
        solicitud.archivo_hc.delete(save=False)
        solicitud.hosvital_estado = "PENDIENTE"
        solicitud.hosvital_mensaje = ""
        solicitud.hosvital_fecha = None
        solicitud.save()
        registrar_log(
            usuario=request.user, accion="Eliminar HC manual",
            modelo="Solicitud", registro_id=solicitud.id,
            detalle=f"Archivo eliminado: {nombre}",
        )
        return Response(self.get_serializer(solicitud).data)

    @action(detail=True, methods=["get"], url_path="descargar-hc")
    def descargar_hc(self, request, pk=None):
        """Descarga el PDF de la HC ya generado."""
        solicitud = self.get_object()
        if not solicitud.archivo_hc:
            return Response(
                {"error": "Esta solicitud no tiene HC generada todavía."},
                status=status.HTTP_404_NOT_FOUND,
            )
        return FileResponse(
            solicitud.archivo_hc.open("rb"),
            as_attachment=True,
            filename=solicitud.archivo_hc.name.split("/")[-1],
            content_type="application/pdf",
        )

    @action(detail=True, methods=["post"], url_path="enviar-hc")
    def enviar_hc(self, request, pk=None):
        """Envía el PDF al paciente por correo o WhatsApp."""
        solicitud = self.get_object()
        if not solicitud.archivo_hc:
            return Response(
                {"error": "Genere primero la HC desde Hosvital."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        canal = (request.data.get("canal") or "").strip().upper()
        destinatario = (request.data.get("destinatario") or "").strip()
        mensaje_extra = request.data.get("mensaje") or ""

        if canal not in ("CORREO", "WHATSAPP"):
            return Response(
                {"error": "Canal inválido. Use CORREO o WHATSAPP."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not destinatario:
            return Response(
                {"error": "Indique el destinatario (correo o número de celular)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        pac = solicitud.paciente
        nombre_pac = f"{pac.nombres} {pac.apellidos}".strip() or "paciente"
        asunto = f"Historia Clínica - {nombre_pac}"
        cuerpo = (
            mensaje_extra
            or f"Adjuntamos la historia clínica solicitada.\n\n"
               f"Paciente: {nombre_pac}\n"
               f"Documento: {pac.tipo_documento} {pac.numero_documento}\n\n"
               f"Clínica Junical — Ventanilla de Historias Clínicas."
        )

        archivo_path = solicitud.archivo_hc.path

        if canal == "CORREO":
            ok, msg = envio.enviar_por_correo(
                destinatario=destinatario, asunto=asunto,
                cuerpo=cuerpo, archivo_path=archivo_path,
            )
        else:  # WHATSAPP
            archivo_url = ""
            try:
                archivo_url = request.build_absolute_uri(solicitud.archivo_hc.url)
            except Exception:
                pass
            nombre_archivo = f"HC_{solicitud.paciente.numero_documento}.pdf"
            ok, msg = envio.enviar_por_whatsapp(
                numero=destinatario,
                mensaje=cuerpo,
                archivo_url=archivo_url,
                archivo_path=archivo_path,
                nombre_archivo=nombre_archivo,
            )

        solicitud.envio_estado = "ENVIADO" if ok else "ERROR"
        solicitud.envio_canal = canal
        solicitud.envio_destinatario = destinatario
        solicitud.envio_mensaje = msg
        solicitud.envio_fecha = timezone.now()
        solicitud.save()

        registrar_log(
            usuario=request.user, accion=f"Enviar HC ({canal})",
            modelo="Solicitud", registro_id=solicitud.id, detalle=msg,
        )

        if not ok:
            return Response({"error": msg}, status=status.HTTP_502_BAD_GATEWAY)

        return Response(self.get_serializer(solicitud).data)
